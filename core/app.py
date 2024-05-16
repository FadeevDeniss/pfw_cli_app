import os

import argparse as ap

from typing import Union, Any
from operator import sub
from datetime import datetime

import pandas as pd
import openpyxl as op

from textwrap import dedent


class PersonalFinancialWallet:
    """
    Приложение командной строки "Персональный финансовый кошелек"

    Позволяет вести учет доходов и расходов и сохраняет информацию
    в файле xlsx
    ...

    Аттрибуты
    __________
    _prog_name : str
        имя приложения для отображения в консоли
    _prog_description : str
        описание приложения для отображения в консоли
    _actions : list
        список команд для приложения
    _root_dir : str
        строка с корневой директорией проекта
    _filename : str
        имя файла для хранения и записи информации
    columns : list
        список наименований для колонок
    categories : list
        список категорий для записей
    db_path : str
        строка содержащая полный путь до файла базы данных
    parser : ArgumentParser
        парсер аргументов командной строки
    """
    _prog_name = 'PFW'
    _prog_description = '''
        Личный Финансовый Кошелек
        ________________________________________________________________
        Приложение командной строки для ведения учета доходов и расходов
    '''

    _balance_table = dedent('''
            __________________________
           | Текущий баланс: %d 
           | ------------------------- 
           | Доходы:         %d          
           | ------------------------- 
           | Расходы:       %d        
           | _________________________ 
        ''')

    _actions = ['balance', 'add', 'search', 'modify']

    _root_dir = os.getcwd()

    _filename = 'db.xlsx'

    _columns = ['Category', 'Amount', 'Date', 'Description', 'Created at', 'Updated at']
    _categories = ['Расход', 'Доход']

    def __init__(self):
        self.db_path = os.path.join(self._root_dir, f'database\\{self._filename}')
        self._create_db_if_not_exists()

        self.parser = ap.ArgumentParser(prog=self._prog_name,
                                        description=dedent(self._prog_description),
                                        formatter_class=ap.RawDescriptionHelpFormatter)

        self.parser.add_argument('action',
                                 choices=self._actions,
                                 help='Выполнить операцию одну на выбор - показать баланс,'
                                      ' добавить редактировать или удалить запись')
        self.parser.add_argument(
            '-c', '--cat', nargs='?', choices=self._categories, dest='category', help='Категория операции'
        )
        self.parser.add_argument('-a', '--amount', dest='amount', help='Сумма')
        self.parser.add_argument('--date', nargs='?', dest='date', help='Дата операции', const=datetime.date)
        self.parser.add_argument('-d', '--desc', nargs='?', dest='description', help='Детали операции')
        self.parser.add_argument('-i', '--index', nargs='?', type=int, dest='idx', help='Индекс записи')

    def count_balance(self, dataframe: pd.DataFrame) -> tuple[int, int, int]:
        """
        Вычисляет баланс, расходы и доходы.

        :param dataframe: дата фрейм с информацией о расходах и доходах
        :return: Кортеж с суммой баланса, общей суммой расходов и доходов
        """
        if dataframe.empty:
            print('\nОшибка: Недостаточно данных для вычисления баланса.'
                  ' Отсутствуют доходы и расходы.')
            raise SystemExit(2)
        income_rows = self._filter_rows(dataframe, {'Category': 'Доход'})
        outcome_rows = self._filter_rows(dataframe, {'Category': 'Расход'})
        income = income_rows.Amount.sum() if not income_rows.empty else 0
        outcome = outcome_rows.Amount.sum() if not outcome_rows.empty else 0
        balance = sub(income, outcome)
        return balance, income, outcome

    def add_record(self, col_values: dict[str, Any]) -> None:
        """
        Добавляет новую запись в таблицу excel.

        :param col_values: Словарь со значениями ячеек,
                           переданных пользователем в аргументах командной строки
        :return: None
        """
        upd_datetime = datetime.now()
        col_values.update({'Created at': upd_datetime})
        self._save_to_excel(col_values)

    def modify_record(self, update_values: dict[str, Any], index: int) -> None:
        """
        Обновляет запись в файле excel.

        :param update_values: Словарь с новыми значениями ячеек
        :param index: Индекс записи для обновления
        :return: None
        """
        update_values['Updated at'] = datetime.now()
        self._save_to_excel(update_values, index)

    def search_record(self, dataframe: pd.DataFrame, conditions: dict[str, str]) -> int:
        """
        Найти запись по значениям ячеек, переданных пользователем в аргументах
        командной строки. Метод меняет дата фрейм "на месте" - т.е. фильтрует
        переданный не создавая при этом новый, и возвращает количество
        найденных строк.

        :param dataframe: датафрейм с информацией о расходах и доходах
        :param conditions: Словарь со значениями ячеек для поиска
        :return: Количество найденных строк
        """
        self._filter_rows(dataframe, conditions, inplace=True)
        rows_count, _ = dataframe.shape
        return rows_count

    def start(self) -> None:
        """
        Запускает парсинг аргументов командной строки, вызывает
        методы класса в зависимости от команды, переданной пользователем и
        выводит информацию в sys.stdout.

        :return: None
        """
        args = self.parser.parse_args()
        _col_values = self.cli_args_to_dict(args)
        if args.action == 'balance':
            dataframe = self._load_dataframe_from_excel()
            print(self._balance_table % self.count_balance(dataframe))
        elif args.action == 'add':
            self.add_record(_col_values)
            print(f'\nДобавлена новая запись:\n\n{self.cell_values_to_string(_col_values.values())}')
        elif args.action == 'search':
            dataframe = self._load_dataframe_from_excel()
            rows_count = self.search_record(dataframe, _col_values)
            print(f'\nНайдено {rows_count} записей\n_________________\n')
            if rows_count > 0:
                print(dataframe)
        else:
            self.modify_record(_col_values, args.idx)
            print(f'\nЗапись обновлена:\n{self.cell_values_to_string(_col_values.values())}')

    def _create_db_if_not_exists(self) -> None:
        """
        Создает файл excel в директории ./database/ если он не создан.

        :return: None
        """
        db_dir_name = os.path.dirname(self.db_path)
        if not os.path.isdir(db_dir_name):
            os.mkdir(db_dir_name)
        if not os.path.isfile(self.db_path):
            wb = op.Workbook()
            sheet = wb.active
            sheet.title = 'Main'
            sheet.append(self._columns)
            wb.save(self.db_path)

    def _load_dataframe_from_excel(self) -> pd.DataFrame:
        """
        Загружает дата фрейм из файла excel в директории ./database/

        :param path:
        :return: Дата фрейм с данными о расходах и доходах из файла excel
        """
        return pd.read_excel(self.db_path, date_format='%Y-%m-%d', parse_dates=['Date'])

    def _save_to_excel(self, values: dict[str, Any], index: int = None) -> None:
        """
        Сохраняет переданные значения в файл excel.

        :param values: Словарь со значениями ячеек
        :param index: Индекс записи (По умолчанию равен None)
        :return: None
        """
        col_indexes = {i[1]: i[0] for i in enumerate(self._columns, 1)}
        workbook = op.load_workbook(self.db_path)
        sheet = workbook['Main']
        index = (sheet.max_row + 1) if index is None else (index + 2)
        for col_name, value in values.items():
            sheet.cell(row=index, column=col_indexes[col_name]).value = value
        workbook.save(self.db_path)

    @staticmethod
    def _filter_rows(dataframe: pd.DataFrame,
                     conditions: dict[str, str],
                     inplace: bool = False) -> Union[pd.DataFrame, None]:
        """
        Фильтрует дата фрейм на основе переданных значений

        :param dataframe: Дата фрейм с данными о расходах и доходах из файла excel
        :param conditions: Значения ячеек для фильтрации
        :param inplace: Опциональный аргумент, определяет создавать ли новый объект дата фрейм,
                        или менять существующий. По умолчанию равен False
        :return: Отфильтрованный дата фрейм или None, если записи не найдены
        """
        expr = ' & '.join([f'{item[0]} == "{item[1]}"' for item in conditions.items()])
        result = dataframe.query(expr, inplace=inplace)
        if not inplace:
            return result

    @staticmethod
    def cli_args_to_dict(args: ap.Namespace) -> dict[str, str]:
        """
        Создает словарь со значениями, переданными пользователем
        в аргументах командной строки

        :param args: Аргументы командной строки
        """
        result = {}
        for attr in ('amount', 'category', 'date', 'description'):
            if getattr(args, attr):
                result[attr.capitalize()] = getattr(args, attr)
        return result

    @staticmethod
    def cell_values_to_string(values: Any):
        """
        Преобразует значения в строковое представление

        :param values: Итерируемый объект
        """
        return " | ".join(str(i) for i in values)
